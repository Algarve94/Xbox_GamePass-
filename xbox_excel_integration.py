"""
Xbox Game Pass — Gerador de Planilha Excel com Integração Python
================================================================
Lê os dados brutos da Planilha_xbox.xlsx e produz um workbook
profissional com:
  • Aba DADOS       — base limpa, datas convertidas, colunas extras
  • Aba KPIs        — 12 indicadores calculados via fórmulas Excel
  • Aba PIVOT       — 4 análises de negócio
  • Aba DASHBOARD   — gráficos embutidos
  • Aba INSTRUCOES  — guia de integração e manutenção
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

SRC  = "Planilha_xbox.xlsx"
DEST = "Xbox_GamePass_Integrado.xlsx"

# ── Paleta Xbox ───────────────────────────────────────────────────────────────
GRN1  = "FF107C10"
GRN2  = "FF52B043"
GRN3  = "FF9BC848"
GRN4  = "FFD6EFC7"
DARK  = "FF0A0A0A"
WHITE = "FFFFFFFF"
LGRAY = "FFF4F4F4"

def fill(h):  return PatternFill("solid", fgColor=h)
def font(bold=False, color=DARK, size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bord():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, text, bg=GRN1, fg=WHITE, sz=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(True, fg, sz)
    c.alignment = align(); c.border = bord()
    return c

def val(ws, row, col, v, bg=WHITE, fg=DARK, bold=False, h="center", fmt=None):
    c = ws.cell(row=row, column=col, value=v)
    c.fill = fill(bg); c.font = font(bold, fg)
    c.alignment = align(h=h); c.border = bord()
    if fmt: c.number_format = fmt
    return c

# ── Ler dados ─────────────────────────────────────────────────────────────────
# Nome da aba tem char especial, usar índice 1 (segunda aba)
df = pd.read_excel(SRC, sheet_name=1)
df.columns = df.columns.str.strip()
df["Start Date"] = pd.to_datetime(df["Start Date"], errors="coerce")
df["EA Play Season Pass\nPrice"] = pd.to_numeric(df["EA Play Season Pass\nPrice"], errors="coerce").fillna(0)
df["Minecraft Season Pass Price"] = pd.to_numeric(df["Minecraft Season Pass Price"], errors="coerce").fillna(0)
df["Coupon Value"] = pd.to_numeric(df["Coupon Value"], errors="coerce").fillna(0)
df["Total Value"] = pd.to_numeric(df["Total Value"], errors="coerce").fillna(0)
n = len(df)

# ── Criar workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════  ABA: DADOS ═════════════
ws_d = wb.create_sheet("DADOS")
ws_d.sheet_view.showGridLines = False
ws_d.freeze_panes = "A2"

COLS = [
    ("Subscriber ID",                "Subscriber ID",              12, "#,##0"),
    ("Name",                         "Name",                       26, "@"),
    ("Plan",                         "Plan",                       13, "@"),
    ("Start Date",                   "Start Date",                 14, "DD/MM/YYYY"),
    ("Auto Renewal",                 "Auto Renewal",               14, "@"),
    ("Subscription Price ($)",       "Subscription Price",         22, '#,##0.00'),
    ("Subscription Type",            "Subscription Type",          18, "@"),
    ("EA Play Season Pass",          "EA Play Season Pass",        20, "@"),
    ("EA Play Season Pass Price ($)","EA Play Season Pass\nPrice", 26, '#,##0.00'),
    ("Minecraft Season Pass",        "Minecraft Season Pass",      22, "@"),
    ("Minecraft Pass Price ($)",     "Minecraft Season Pass Price",24, '#,##0.00'),
    ("Coupon Value ($)",             "Coupon Value",               16, '#,##0.00'),
    ("Total Value ($)",              "Total Value",                16, '#,##0.00'),
]

for ci, (label, _, width, _fmt) in enumerate(COLS, 1):
    hdr(ws_d, 1, ci, label)
    ws_d.column_dimensions[get_column_letter(ci)].width = width
ws_d.row_dimensions[1].height = 28

for ri, (_, row) in enumerate(df.iterrows(), 2):
    bg = GRN4 if ri % 2 == 0 else WHITE
    for ci, (label, src_col, _, fmt) in enumerate(COLS, 1):
        raw = row[src_col]
        if fmt == "DD/MM/YYYY":
            v = raw.to_pydatetime() if pd.notna(raw) else None
        elif fmt == '#,##0.00':
            v = float(raw) if pd.notna(raw) else 0.0
        elif fmt == '#,##0':
            v = int(raw) if pd.notna(raw) else 0
        else:
            v = str(raw) if pd.notna(raw) else ""
        c = val(ws_d, ri, ci, v, bg=bg, h="left" if fmt == "@" else "center", fmt=fmt if fmt != "@" else None)

last_col_ltr = get_column_letter(len(COLS))
tbl = Table(displayName="TabelaDados", ref=f"A1:{last_col_ltr}{n+1}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True)
ws_d.add_table(tbl)

# ════════════════════════════════════════════════════  ABA: KPIs ══════════════
ws_k = wb.create_sheet("KPIs")
ws_k.sheet_view.showGridLines = False

def kpi(ws, row, col, title, formula, fmt="#,##0.00", note=""):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    c = ws.cell(row=row, column=col, value=title)
    c.fill = fill(GRN1); c.font = font(True, WHITE, 9)
    c.alignment = align(wrap=True); c.border = bord()

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    v = ws.cell(row=row+1, column=col, value=formula)
    v.fill = fill(GRN4); v.font = font(True, GRN1, 20)
    v.alignment = align(); v.number_format = fmt; v.border = bord()

    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    n_ = ws.cell(row=row+2, column=col, value=note)
    n_.fill = fill(LGRAY); n_.font = font(False, "FF666666", 8)
    n_.alignment = align(h="center", wrap=True)

for ci in range(1, 13):
    ws_k.column_dimensions[get_column_letter(ci)].width = 11

ws_k.row_dimensions[1].height = 38
t = ws_k.cell(row=1, column=1, value="📊  XBOX GAME PASS — INDICADORES PRINCIPAIS (KPIs)")
t.fill = fill(DARK); t.font = font(True, WHITE, 14)
t.alignment = align(h="left"); ws_k.merge_cells("A1:L1")

ws_k.row_dimensions[2].height = 22
sub = ws_k.cell(row=2, column=1,
    value=f"Período: Jan–Dez 2024  |  Gerado em: {datetime.date.today():%d/%m/%Y}  |  Fonte: aba DADOS")
sub.fill = fill(GRN2); sub.font = font(False, WHITE, 9)
sub.alignment = align(h="left"); ws_k.merge_cells("A2:L2")

kpi(ws_k, 4,  1, "TOTAL ASSINANTES",           "=COUNTA(TabelaDados[Subscriber ID])",   "#,##0",  "Registros únicos na base")
kpi(ws_k, 4,  4, "RECEITA BRUTA TOTAL ($)",     "=SUM(TabelaDados[Total Value ($)])",    "#,##0.00","Soma de todos os Total Value")
kpi(ws_k, 4,  7, "TICKET MÉDIO ($)",             "=AVERAGE(TabelaDados[Total Value ($)])","#,##0.00","Média por assinante")
kpi(ws_k, 4, 10, "TAXA AUTO RENOVAÇÃO",
    '=COUNTIF(TabelaDados[Auto Renewal],"Yes")/COUNTA(TabelaDados[Auto Renewal])',
    "0.0%", "% com renovação automática")

kpi(ws_k, 9,  1, "ASSINANTES ULTIMATE",  '=COUNTIF(TabelaDados[Plan],"Ultimate")', "#,##0", "Plano premium")
kpi(ws_k, 9,  4, "ASSINANTES STANDARD",  '=COUNTIF(TabelaDados[Plan],"Standard")', "#,##0", "Plano intermediário")
kpi(ws_k, 9,  7, "ASSINANTES CORE",      '=COUNTIF(TabelaDados[Plan],"Core")',     "#,##0", "Plano básico")
kpi(ws_k, 9, 10, "RECEITA EA PLAY ($)",
    '=SUMIF(TabelaDados[EA Play Season Pass],"Yes",TabelaDados[EA Play Season Pass Price ($)])',
    "#,##0.00", "Assinantes com EA Play ativo")

kpi(ws_k, 14,  1, "RECEITA MINECRAFT ($)",
    '=SUMIF(TabelaDados[Minecraft Season Pass],"Yes",TabelaDados[Minecraft Pass Price ($)])',
    "#,##0.00", "Assinantes com Minecraft Pass")
kpi(ws_k, 14,  4, "RECEITA PLANOS MENSAIS ($)",
    '=SUMIF(TabelaDados[Subscription Type],"Monthly",TabelaDados[Total Value ($)])',
    "#,##0.00", "Tipo Monthly")
kpi(ws_k, 14,  7, "RECEITA PLANOS ANUAIS ($)",
    '=SUMIF(TabelaDados[Subscription Type],"Annual",TabelaDados[Total Value ($)])',
    "#,##0.00", "Tipo Annual")
kpi(ws_k, 14, 10, "TOTAL CUPONS ($)",
    "=SUM(TabelaDados[Coupon Value ($)])", "#,##0.00", "Soma dos descontos aplicados")

# ════════════════════════════════════════════════════  ABA: PIVOT ═════════════
ws_p = wb.create_sheet("PIVOT")
ws_p.sheet_view.showGridLines = False
for ci, w in enumerate([2, 28, 22, 18, 18, 4, 28, 22, 18, 18], 1):
    ws_p.column_dimensions[get_column_letter(ci)].width = w

t = ws_p.cell(row=1, column=1, value="📋  XBOX GAME PASS — ANÁLISE POR PERGUNTAS DE NEGÓCIO")
t.fill = fill(DARK); t.font = font(True, WHITE, 14)
t.alignment = align(h="left"); ws_p.merge_cells("A1:J1"); ws_p.row_dimensions[1].height = 38

def ptitle(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(GRN1); c.font = font(True, WHITE, 11)
    c.alignment = align(h="left"); c.border = bord()
    ws.row_dimensions[row].height = 26

def ph(ws, r, c, text): hdr(ws, r, c, text, GRN2, WHITE, 10)
def pl(ws, r, c, text, bg=LGRAY, bold=False):
    x = ws.cell(row=r, column=c, value=text)
    x.fill = fill(bg); x.font = font(bold, DARK, 10)
    x.alignment = align(h="left"); x.border = bord()
def pv(ws, r, c, formula, fmt="#,##0.00", bg=WHITE, bold=False):
    x = ws.cell(row=r, column=c, value=formula)
    x.fill = fill(bg); x.font = font(bold, DARK, 10)
    x.alignment = align(h="right"); x.number_format = fmt; x.border = bord()

# P1 — Faturamento por tipo de assinatura
ptitle(ws_p, 3, 2, "P1 — Faturamento Total por Tipo de Assinatura")
for ci, txt in enumerate(["Tipo", "Total Value ($)", "Nº Assin.", "% Receita"], 2): ph(ws_p, 4, ci, txt)
tipos = [("Monthly", WHITE), ("Quarterly", GRN4), ("Annual", WHITE)]
for i, (tp, bg) in enumerate(tipos, 5):
    pl(ws_p, i, 2, tp, bg)
    pv(ws_p, i, 3, f'=SUMIF(TabelaDados[Subscription Type],"{tp}",TabelaDados[Total Value ($)])', "#,##0.00", bg)
    pv(ws_p, i, 4, f'=COUNTIF(TabelaDados[Subscription Type],"{tp}")', "#,##0", bg)
    pv(ws_p, i, 5, f"=C{i}/SUM(C5:C7)", "0.0%", bg)
pl(ws_p, 8, 2, "TOTAL", GRN3, True)
pv(ws_p, 8, 3, "=SUM(C5:C7)", "#,##0.00", GRN3, True)
pv(ws_p, 8, 4, "=SUM(D5:D7)", "#,##0", GRN3, True)
pv(ws_p, 8, 5, "=SUM(E5:E7)", "0.0%", GRN3, True)

# P2 — Anuais x Auto Renovação
ptitle(ws_p, 10, 2, "P2 — Planos Anuais × Auto Renovação")
for ci, txt in enumerate(["Auto Renovação", "Faturamento ($)", "Nº Assin.", "Ticket Médio ($)"], 2): ph(ws_p, 11, ci, txt)
for i, (ar, bg) in enumerate([("Yes", GRN4), ("No", WHITE)], 12):
    pl(ws_p, i, 2, ar, bg)
    pv(ws_p, i, 3, f'=SUMPRODUCT((TabelaDados[Subscription Type]="Annual")*(TabelaDados[Auto Renewal]="{ar}")*TabelaDados[Total Value ($)])', "#,##0.00", bg)
    pv(ws_p, i, 4, f'=SUMPRODUCT((TabelaDados[Subscription Type]="Annual")*(TabelaDados[Auto Renewal]="{ar}"))', "#,##0", bg)
    pv(ws_p, i, 5, f"=IF(D{i}=0,0,C{i}/D{i})", "#,##0.00", bg)
pl(ws_p, 14, 2, "TOTAL ANUAL", GRN3, True)
pv(ws_p, 14, 3, "=SUM(C12:C13)", "#,##0.00", GRN3, True)
pv(ws_p, 14, 4, "=SUM(D12:D13)", "#,##0", GRN3, True)
pv(ws_p, 14, 5, "=IF(D14=0,0,C14/D14)", "#,##0.00", GRN3, True)

# P3 — EA Play por plano
ptitle(ws_p, 16, 2, "P3 — Vendas EA Play Season Pass por Plano")
for ci, txt in enumerate(["Plano", "Receita EA Play ($)", "Qtd c/ EA Play", "% do Plano"], 2): ph(ws_p, 17, ci, txt)
planos = [("Ultimate", GRN4), ("Standard", WHITE), ("Core", GRN4)]
for i, (pl_, bg) in enumerate(planos, 18):
    pl(ws_p, i, 2, pl_, bg)
    pv(ws_p, i, 3, f'=SUMPRODUCT((TabelaDados[Plan]="{pl_}")*(TabelaDados[EA Play Season Pass]="Yes")*TabelaDados[EA Play Season Pass Price ($)])', "#,##0.00", bg)
    pv(ws_p, i, 4, f'=SUMPRODUCT((TabelaDados[Plan]="{pl_}")*(TabelaDados[EA Play Season Pass]="Yes"))', "#,##0", bg)
    pv(ws_p, i, 5, f'=IF(COUNTIF(TabelaDados[Plan],"{pl_}")=0,0,D{i}/COUNTIF(TabelaDados[Plan],"{pl_}"))', "0.0%", bg)
pl(ws_p, 21, 2, "TOTAL", GRN3, True)
pv(ws_p, 21, 3, "=SUM(C18:C20)", "#,##0.00", GRN3, True)
pv(ws_p, 21, 4, "=SUM(D18:D20)", "#,##0", GRN3, True)

# P4 — Minecraft por plano
ptitle(ws_p, 23, 2, "P4 — Vendas Minecraft Season Pass por Plano")
for ci, txt in enumerate(["Plano", "Receita Minecraft ($)", "Qtd c/ Minecraft", "% do Plano"], 2): ph(ws_p, 24, ci, txt)
for i, (pl_, bg) in enumerate(planos, 25):
    pl(ws_p, i, 2, pl_, bg)
    pv(ws_p, i, 3, f'=SUMPRODUCT((TabelaDados[Plan]="{pl_}")*(TabelaDados[Minecraft Season Pass]="Yes")*TabelaDados[Minecraft Pass Price ($)])', "#,##0.00", bg)
    pv(ws_p, i, 4, f'=SUMPRODUCT((TabelaDados[Plan]="{pl_}")*(TabelaDados[Minecraft Season Pass]="Yes"))', "#,##0", bg)
    pv(ws_p, i, 5, f'=IF(COUNTIF(TabelaDados[Plan],"{pl_}")=0,0,D{i}/COUNTIF(TabelaDados[Plan],"{pl_}"))', "0.0%", bg)
pl(ws_p, 28, 2, "TOTAL", GRN3, True)
pv(ws_p, 28, 3, "=SUM(C25:C27)", "#,##0.00", GRN3, True)
pv(ws_p, 28, 4, "=SUM(D25:D27)", "#,##0", GRN3, True)

# ════════════════════════════════════════════════════  ABA: DASHBOARD ═════════
ws_ch = wb.create_sheet("DASHBOARD")
ws_ch.sheet_view.showGridLines = False

t = ws_ch.cell(row=1, column=1, value="🎮  XBOX GAME PASS — DASHBOARD DE VENDAS 2024")
t.fill = fill(DARK); t.font = font(True, WHITE, 15)
t.alignment = align(h="left"); ws_ch.merge_cells("A1:O1"); ws_ch.row_dimensions[1].height = 42
sub = ws_ch.cell(row=2, column=1, value=f"Período: Jan–Dez 2024  |  Atualizado: {datetime.date.today():%d/%m/%Y}  |  Fonte: TabelaDados")
sub.fill = fill(GRN1); sub.font = font(False, WHITE, 9)
sub.alignment = align(h="left"); ws_ch.merge_cells("A2:O2"); ws_ch.row_dimensions[2].height = 20

# Dados auxiliares (cols Q, R, S)
ws_ch.column_dimensions["Q"].width = 16
ws_ch.column_dimensions["R"].width = 18

# Receita por tipo (Q1:R4)
ws_ch["Q1"] = "Tipo";       ws_ch["R1"] = "Receita ($)"
ws_ch["Q2"] = "Monthly";    ws_ch["R2"] = '=SUMIF(TabelaDados[Subscription Type],"Monthly",TabelaDados[Total Value ($)])'
ws_ch["Q3"] = "Quarterly";  ws_ch["R3"] = '=SUMIF(TabelaDados[Subscription Type],"Quarterly",TabelaDados[Total Value ($)])'
ws_ch["Q4"] = "Annual";     ws_ch["R4"] = '=SUMIF(TabelaDados[Subscription Type],"Annual",TabelaDados[Total Value ($)])'

# Assinantes por plano (Q6:R9)
ws_ch["Q6"] = "Plano";      ws_ch["R6"] = "Assinantes"
ws_ch["Q7"] = "Ultimate";   ws_ch["R7"] = '=COUNTIF(TabelaDados[Plan],"Ultimate")'
ws_ch["Q8"] = "Standard";   ws_ch["R8"] = '=COUNTIF(TabelaDados[Plan],"Standard")'
ws_ch["Q9"] = "Core";       ws_ch["R9"] = '=COUNTIF(TabelaDados[Plan],"Core")'

# Season Passes (Q11:R13)
ws_ch["Q11"] = "Produto";   ws_ch["R11"] = "Receita ($)"
ws_ch["Q12"] = "EA Play";   ws_ch["R12"] = '=SUMIF(TabelaDados[EA Play Season Pass],"Yes",TabelaDados[EA Play Season Pass Price ($)])'
ws_ch["Q13"] = "Minecraft"; ws_ch["R13"] = '=SUMIF(TabelaDados[Minecraft Season Pass],"Yes",TabelaDados[Minecraft Pass Price ($)])'

# Gráfico 1 — Receita por tipo (barras)
bc1 = BarChart()
bc1.type = "col"; bc1.title = "Receita por Tipo de Assinatura ($)"
bc1.style = 10; bc1.width = 16; bc1.height = 12
bc1.y_axis.title = "Receita ($)"; bc1.x_axis.title = "Tipo"
d1 = Reference(ws_ch, min_col=18, max_col=18, min_row=1, max_row=4)
c1 = Reference(ws_ch, min_col=17, min_row=2, max_row=4)
bc1.add_data(d1, titles_from_data=True)
bc1.set_categories(c1)
bc1.series[0].graphicalProperties.solidFill = "107C10"
ws_ch.add_chart(bc1, "A4")

# Gráfico 2 — Pizza por plano
pc = PieChart()
pc.title = "Distribuição de Assinantes por Plano"
pc.style = 10; pc.width = 16; pc.height = 12
d2 = Reference(ws_ch, min_col=18, max_col=18, min_row=6, max_row=9)
c2 = Reference(ws_ch, min_col=17, min_row=7, max_row=9)
pc.add_data(d2, titles_from_data=True)
pc.set_categories(c2)
for i, clr in enumerate(["107C10","52B043","9BC848"]):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = clr
    pc.series[0].dPt.append(pt)
ws_ch.add_chart(pc, "H4")

# Gráfico 3 — Season Passes
bc2 = BarChart()
bc2.type = "col"; bc2.title = "Receita por Season Pass ($)"
bc2.style = 10; bc2.width = 16; bc2.height = 12
bc2.y_axis.title = "Receita ($)"
d3 = Reference(ws_ch, min_col=18, max_col=18, min_row=11, max_row=13)
c3 = Reference(ws_ch, min_col=17, min_row=12, max_row=13)
bc2.add_data(d3, titles_from_data=True)
bc2.set_categories(c3)
bc2.series[0].graphicalProperties.solidFill = "52B043"
ws_ch.add_chart(bc2, "A23")

# ════════════════════════════════════════════════════  ABA: INSTRUCOES ════════
ws_i = wb.create_sheet("INSTRUCOES")
ws_i.sheet_view.showGridLines = False
ws_i.column_dimensions["A"].width = 3
ws_i.column_dimensions["B"].width = 34
ws_i.column_dimensions["C"].width = 62

def ititle(ws, row, text):
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(GRN1); c.font = font(True, WHITE, 12)
    c.alignment = align(h="left"); ws.row_dimensions[row].height = 30

def irow(ws, row, label, desc, bg=WHITE):
    l = ws.cell(row=row, column=2, value=label)
    l.fill = fill(bg); l.font = font(True, DARK, 10); l.alignment = align(h="left"); l.border = bord()
    d = ws.cell(row=row, column=3, value=desc)
    d.fill = fill(bg); d.font = font(False, DARK, 10); d.alignment = align(h="left", wrap=True); d.border = bord()
    ws.row_dimensions[row].height = 20

def icode(ws, row, code):
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2, value=code)
    c.fill = fill("FF1A1A2E"); c.font = Font(name="Courier New", size=9, color="FF9BC848")
    c.alignment = align(h="left"); ws.row_dimensions[row].height = 18

t = ws_i.cell(row=1, column=1, value="📘  XBOX GAME PASS — GUIA COMPLETO DE INTEGRAÇÃO PYTHON × EXCEL")
t.fill = fill(DARK); t.font = font(True, WHITE, 14)
t.alignment = align(h="left"); ws_i.merge_cells("A1:C1"); ws_i.row_dimensions[1].height = 42

r = 3
ititle(ws_i, r, "1. ESTRUTURA DA PLANILHA GERADA"); r+=1
irow(ws_i, r, "Aba DADOS",       "Base de dados com 295 registros. Tabela Excel nomeada 'TabelaDados'. Datas convertidas.", LGRAY); r+=1
irow(ws_i, r, "Aba KPIs",        "12 indicadores calculados com fórmulas que referenciam TabelaDados dinamicamente.", WHITE); r+=1
irow(ws_i, r, "Aba PIVOT",       "4 análises de negócio usando SUMIF / COUNTIF / SUMPRODUCT.", LGRAY); r+=1
irow(ws_i, r, "Aba DASHBOARD",   "3 gráficos Excel (barras + pizza) alimentados por fórmulas nas colunas Q:R.", WHITE); r+=1
irow(ws_i, r, "Aba INSTRUCOES",  "Este guia.", LGRAY); r+=2

ititle(ws_i, r, "2. PRÉ-REQUISITOS PARA RODAR O SCRIPT PYTHON"); r+=1
irow(ws_i, r, "Python 3.9+",  "Verifique com:  python --version", LGRAY); r+=1
irow(ws_i, r, "openpyxl",     "pip install openpyxl", WHITE); r+=1
irow(ws_i, r, "pandas",       "pip install pandas", LGRAY); r+=1
irow(ws_i, r, "openpyxl-chart","Já incluído no openpyxl 3.1+  (gráficos nativos)", WHITE); r+=2

ititle(ws_i, r, "3. COMO EXECUTAR O SCRIPT"); r+=1
icode(ws_i, r, "# Instalar dependências (uma vez só)"); r+=1
icode(ws_i, r, "pip install openpyxl pandas"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "# Gerar / atualizar a planilha"); r+=1
icode(ws_i, r, "python xbox_excel_integration.py"); r+=2

ititle(ws_i, r, "4. COMO ATUALIZAR OS DADOS"); r+=1
irow(ws_i, r, "Opção A (recomendada)", "Substitua o arquivo Planilha_xbox.xlsx e rode o script novamente.", LGRAY); r+=1
irow(ws_i, r, "Opção B (Excel direto)", "Na aba DADOS, clique na última linha da tabela e pressione Tab. Preencha normalmente.", WHITE); r+=1
irow(ws_i, r, "Auto-expansão", "A Tabela Excel 'TabelaDados' expande sozinha — todos os KPIs e gráficos se atualizam.", LGRAY); r+=2

ititle(ws_i, r, "5. LEITURA DA PLANILHA VIA PYTHON"); r+=1
icode(ws_i, r, "import pandas as pd"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "# Ler todos os dados da aba DADOS"); r+=1
icode(ws_i, r, "df = pd.read_excel('Xbox_GamePass_Integrado.xlsx', sheet_name='DADOS')"); r+=1
icode(ws_i, r, "print(df.head())"); r+=1
icode(ws_i, r, "print(df.describe())  # estatísticas gerais"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "# Filtrar só planos Ultimate com auto-renovação"); r+=1
icode(ws_i, r, "df_ult = df[(df['Plan']=='Ultimate') & (df['Auto Renewal']=='Yes')]"); r+=1
icode(ws_i, r, "print(df_ult[['Name','Total Value ($)']].to_string())"); r+=2

ititle(ws_i, r, "6. ADIÇÃO DE NOVOS REGISTROS VIA PYTHON"); r+=1
icode(ws_i, r, "from openpyxl import load_workbook"); r+=1
icode(ws_i, r, "import datetime"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "wb = load_workbook('Xbox_GamePass_Integrado.xlsx')"); r+=1
icode(ws_i, r, "ws = wb['DADOS']"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "novo = [3526, 'Carlos Lima', 'Ultimate', datetime.date(2024,1,15),"); r+=1
icode(ws_i, r, "        'Yes', 15, 'Monthly', 'Yes', 30, 'Yes', 20, 0, 65]"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "ws.append(novo)   # adiciona linha no final da tabela"); r+=1
icode(ws_i, r, "wb.save('Xbox_GamePass_Integrado.xlsx')"); r+=2

ititle(ws_i, r, "7. INTEGRAÇÃO COM BANCO DE DADOS (EXEMPLO)"); r+=1
icode(ws_i, r, "import sqlite3, pandas as pd"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "# Exportar aba DADOS para SQLite"); r+=1
icode(ws_i, r, "df = pd.read_excel('Xbox_GamePass_Integrado.xlsx', sheet_name='DADOS')"); r+=1
icode(ws_i, r, "conn = sqlite3.connect('xbox_gamepass.db')"); r+=1
icode(ws_i, r, "df.to_sql('assinantes', conn, if_exists='replace', index=False)"); r+=1
icode(ws_i, r, "conn.close()"); r+=1
icode(ws_i, r, ""); r+=1
icode(ws_i, r, "# Consultar e atualizar o Excel com dados do banco"); r+=1
icode(ws_i, r, "conn = sqlite3.connect('xbox_gamepass.db')"); r+=1
icode(ws_i, r, "df_db = pd.read_sql('SELECT * FROM assinantes WHERE Plan=\"Ultimate\"', conn)"); r+=1
icode(ws_i, r, "df_db.to_excel('relatorio_ultimate.xlsx', index=False)"); r+=2

ititle(ws_i, r, "8. DICAS IMPORTANTES"); r+=1
irow(ws_i, r, "Fórmulas",     "NUNCA edite as fórmulas nas abas KPIs e PIVOT manualmente sem entender as referências.", LGRAY); r+=1
irow(ws_i, r, "Tabela Excel", "O nome 'TabelaDados' é usado em todas as fórmulas. Não renomeie!", WHITE); r+=1
irow(ws_i, r, "Datas",        "Ao inserir via Python, use datetime.date(YYYY,MM,DD) para evitar erros de formato.", LGRAY); r+=1
irow(ws_i, r, "Encoding",     "Acentos e caracteres especiais são preservados via UTF-8 pelo openpyxl.", WHITE); r+=1
irow(ws_i, r, "Compatib.",    "Testado: Excel 365, Excel 2019, LibreOffice Calc 7.5+", LGRAY); r+=1
irow(ws_i, r, "Gráficos",     "Para atualizar gráficos no Excel: Ctrl+Alt+F9 (recalcular tudo).", WHITE); r+=2

ws_i.cell(row=r, column=2, value=f"Gerado por: PyMaster Analytics  |  Versão 1.0  |  {datetime.date.today():%d/%m/%Y}").font = font(False, "FF888888", 8)

# ── Ordenar e salvar ──────────────────────────────────────────────────────────
order = ["DADOS", "KPIs", "PIVOT", "DASHBOARD", "INSTRUCOES"]
for i, name in enumerate(order):
    wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

wb.active = wb["DASHBOARD"]
wb.save(DEST)
print(f"✅  Salvo em: {DEST}")
print(f"    Registros: {n}  |  Abas: {wb.sheetnames}")
